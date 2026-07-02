"""
database.py
-----------
Handles all reading/writing to the Excel workbook that acts as the
system's persistent database. No business rules live here — this
module only knows how to move data in and out of bank_data.xlsx.
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ACCOUNTS_SHEET = "Accounts"
TRANSACTIONS_SHEET = "Transactions"

ACCOUNT_HEADERS = [
    "Account Number", "Name", "Age", "Phone Numbers",
    "Account Type", "Balance", "Status", "Created On"
]

TRANSACTION_HEADERS = [
    "Transaction ID", "Account Number", "Date/Time",
    "Type", "Amount", "Balance After", "Description"
]


class ExcelDatabase:
    """Thin wrapper around an .xlsx file used as the storage engine."""

    def __init__(self, filepath: str = "bank_data.xlsx"):
        self.filepath = filepath
        if not os.path.exists(self.filepath):
            self._create_new_workbook()

    # ---------- setup ----------

    def _create_new_workbook(self):
        wb = Workbook()
        accounts_ws = wb.active
        accounts_ws.title = ACCOUNTS_SHEET
        accounts_ws.append(ACCOUNT_HEADERS)

        txn_ws = wb.create_sheet(TRANSACTIONS_SHEET)
        txn_ws.append(TRANSACTION_HEADERS)

        wb.save(self.filepath)

    def _load(self):
        return load_workbook(self.filepath)

    # ---------- accounts ----------

    def get_all_accounts(self) -> list[dict]:
        wb = self._load()
        ws: Worksheet = wb[ACCOUNTS_SHEET]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            rows.append(dict(zip(ACCOUNT_HEADERS, row)))
        return rows

    def get_account(self, account_number: int) -> dict | None:
        for row in self.get_all_accounts():
            if int(row["Account Number"]) == int(account_number):
                return row
        return None

    def insert_account(self, account: dict):
        wb = self._load()
        ws = wb[ACCOUNTS_SHEET]
        ws.append([account[h] for h in ACCOUNT_HEADERS])
        wb.save(self.filepath)

    def update_account(self, account_number: int, **fields):
        wb = self._load()
        ws = wb[ACCOUNTS_SHEET]
        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None and int(row[0].value) == int(account_number):
                for col_idx, header in enumerate(ACCOUNT_HEADERS):
                    if header in fields:
                        row[col_idx].value = fields[header]
                wb.save(self.filepath)
                return True
        return False

    def delete_account(self, account_number: int):
        wb = self._load()
        ws = wb[ACCOUNTS_SHEET]
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=1)
            if cell.value is not None and int(cell.value) == int(account_number):
                ws.delete_rows(row_idx)
                wb.save(self.filepath)
                return True
        return False

    def account_number_exists(self, account_number: int) -> bool:
        return self.get_account(account_number) is not None

    # ---------- transactions ----------

    def add_transaction(self, account_number: int, txn_type: str,
                         amount: float, balance_after: float, description: str = ""):
        wb = self._load()
        ws = wb[TRANSACTIONS_SHEET]
        txn_id = ws.max_row  # header row = row 1, so this gives a simple incrementing id
        ws.append([
            txn_id,
            account_number,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            txn_type,
            amount,
            balance_after,
            description
        ])
        wb.save(self.filepath)

    def get_transactions(self, account_number: int) -> list[dict]:
        wb = self._load()
        ws: Worksheet = wb[TRANSACTIONS_SHEET]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] is not None and int(row[1]) == int(account_number):
                rows.append(dict(zip(TRANSACTION_HEADERS, row)))
        return rows